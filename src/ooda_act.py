import json
import os
from collections import Counter
from datetime import datetime
from typing import Dict, List

import yaml
from dotenv import load_dotenv
from openai import OpenAI
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

from db import get_conn, is_remote

MODEL = "gpt-5-mini"  # come hai scelto tu


def get_brand(cfg: Dict) -> str:
	return os.getenv("BRAND", cfg.get("project", {}).get("brand", "")).strip()


def ensure_act_table(conn) -> None:
	cur = conn.cursor()
	cur.execute("""
	CREATE TABLE IF NOT EXISTS runs_act (
		id INTEGER PRIMARY KEY AUTOINCREMENT,
		brand TEXT,
		act_json TEXT,
		created_at TEXT NOT NULL DEFAULT (datetime('now'))
	)
	""")
	conn.commit()


def table_exists(conn, table_name: str) -> bool:
	cur = conn.cursor()
	if is_remote():
		cur.execute("""
			SELECT 1
			FROM information_schema.tables
			WHERE table_schema = 'public' AND table_name = %s
			LIMIT 1
		""", (table_name,))
	else:
		cur.execute("""
			SELECT name FROM sqlite_master
			WHERE type='table' AND name=?
		""", (table_name,))
	return cur.fetchone() is not None


def fetch_full_ooda_view(db_path: str, brand: str, limit: int = 50) -> List[Dict]:
	"""
	Join completo: raw + orient + decide.
	"""
	conn = get_conn(db_path)

	if not table_exists(conn, "items_decide"):
		conn.close()
		raise RuntimeError("Table items_decide not found. Run: python src/ooda_decide.py")

	cur = conn.cursor()
	cur.execute("""
		SELECT
			d.id AS decide_id,
			d.created_at AS decided_at,
			d.decide_json AS decide_json,
			d.orient_id AS orient_id,
			d.raw_item_id AS raw_item_id,

			r.source AS source,
			r.title AS title,
			r.url AS url,
			r.content AS snippet,
			r.metadata_json AS raw_metadata_json,
			r.created_at AS observed_at,
			r.published_at AS published_at,

			o.orient_json AS orient_json,
			o.created_at AS oriented_at

		FROM items_decide d
		LEFT JOIN items_raw r ON r.id = d.raw_item_id
		LEFT JOIN items_orient o ON o.id = d.orient_id
		WHERE r.published_at >= datetime('now','-7 days')
		AND d.brand = ?
		ORDER BY d.id DESC
		LIMIT ?
	""", (brand, limit))

	rows = []
	for r in cur.fetchall():
		decide = json.loads(r["decide_json"]) if r["decide_json"] else {}
		orient = {}
		try:
			orient = json.loads(r["orient_json"]) if r["orient_json"] else {}
		except Exception:
			orient = {}

		rows.append({
			"decide_id": r["decide_id"],
			"decided_at": r["decided_at"],
			"orient_id": r["orient_id"],
			"raw_item_id": r["raw_item_id"],

			"source": r["source"],
			"title": r["title"],
			"url": r["url"],
			"snippet": (r["snippet"] or "")[:800].replace("\n", " ").strip(),
			"observed_at": r["observed_at"],
			"published_at": r["published_at"],

			# ORIENT fields
			"claim_summary": orient.get("claim_summary"),
			"narrative_category": orient.get("narrative_category"),
			"reputational_risk": orient.get("reputational_risk"),
			"severity": orient.get("severity"),
			"confidence": orient.get("confidence"),
			"verification_steps": orient.get("verification_steps", []),

			# DECIDE fields
			"intent_framing": decide.get("intent_framing"),
			"urgency": decide.get("urgency"),
			"recommended_action": decide.get("recommended_action"),
			"escalation_team": decide.get("escalation_team", []),
			"rationale": decide.get("rationale"),
			"no_regret_move": decide.get("no_regret_move"),

			# timestamps
			"oriented_at": r["oriented_at"],
		})

	conn.close()
	return rows


def compute_stats(items: List[Dict]) -> Dict:
	intents = [str(i.get("intent_framing") or "NEUTRAL").upper() for i in items]
	urgencies = [str(i.get("urgency") or "low").lower() for i in items]
	cats = [str(i.get("narrative_category") or "other") for i in items]
	risks = [str(i.get("reputational_risk") or "low") for i in items]
	severities = [i.get("severity") for i in items if isinstance(i.get("severity"), (int, float))]

	intent_counts = Counter(intents)
	urgency_counts = Counter(urgencies)
	cat_counts = Counter(cats)
	risk_counts = Counter(risks)

	severity_stats = {}
	if severities:
		severity_stats = {
			"min": int(min(severities)),
			"max": int(max(severities)),
			"avg": round(sum(severities) / len(severities), 2),
		}
	else:
		severity_stats = {"min": None, "max": None, "avg": None}

	# top 5 items by severity
	top_by_severity = sorted(
		[i for i in items if isinstance(i.get("severity"), (int, float))],
		key=lambda x: x["severity"],
		reverse=True
	)[:5]

	# Priority heuristic: THREAT high/medium urgency OR severity>=60
	priority = []
	for it in items:
		intent = (it.get("intent_framing") or "NEUTRAL").upper()
		urg = (it.get("urgency") or "low").lower()
		sev = it.get("severity")
		if intent == "THREAT" and urg in ["high", "medium"]:
			priority.append(it)
		elif isinstance(sev, (int, float)) and sev >= 60:
			priority.append(it)

	return {
		"counts": {
			"items_total": len(items),
			"intent_distribution": dict(intent_counts),
			"urgency_distribution": dict(urgency_counts),
			"category_distribution": dict(cat_counts),
			"reputational_risk_distribution": dict(risk_counts),
		},
		"severity_stats": severity_stats,
		"top_by_severity": [
			{
				"title": x.get("title"),
				"url": x.get("url"),
				"severity": x.get("severity"),
				"intent_framing": x.get("intent_framing"),
				"recommended_action": x.get("recommended_action"),
			} for x in top_by_severity
		],
		"priority_candidates_count": len(priority),
	}


def build_act_prompt(brand: str, stats: Dict, items: List[Dict]) -> str:
	"""
	ACT aggregato “full”: Executive brief + action plan + monitoring + comms + triggers.
	Passa:
	- stats
	- top by severity
	- subset di items (non tutti per token)
	"""
	# compressione items per prompt: prendiamo max 12
	compact_items = []
	for it in items[:12]:
		compact_items.append({
			"title": it.get("title"),
			"url": it.get("url"),
			"snippet": it.get("snippet"),
			"severity": it.get("severity"),
			"narrative_category": it.get("narrative_category"),
			"reputational_risk": it.get("reputational_risk"),
			"claim_summary": it.get("claim_summary"),
			"intent_framing": it.get("intent_framing"),
			"urgency": it.get("urgency"),
			"recommended_action": it.get("recommended_action"),
			"no_regret_move": it.get("no_regret_move"),
		})

	payload = {
		"brand": brand,
		"stats": stats,
		"items_sample": compact_items,
		"rules": {
			"THREAT": "containment + fact-check + alignment PR/Legal + rapid response plan",
			"DEFENSE": "do NOT escalate; monitor + reputation reinforcement (brand fighting issue)",
			"OPPORTUNITY": "leverage positive narrative",
			"NEUTRAL/NOISE": "log/ignore unless triggers fire"
		}
	}

	return f"""
ROLE: You are the ACT module of an OODA Loop AI early-warning system for brand reputation monitoring.
You are an executive coordinator producing an action package.

FRAMEWORK CONSTRAINT (ACT ONLY):
- This is ACT: convert prior decisions into an executable plan for the next 4 hours.
- Do NOT reclassify items. Do NOT invent new facts. Use ONLY the input JSON provided (stats + items_sample + rules).
- If information is insufficient, explicitly choose conservative 'no-regret' actions and monitoring.

BRAND: {brand}

OBJECTIVE:
Produce ONE aggregated action package for the next 4 hours, grounded in the input data and consistent with ORIENT+DECIDE outputs.

CRITICAL INTERPRETATION RULE:
- If the item describes enforcement already happening (seizure/crackdown/counterfeit removal) and the brand is not accused, treat as DEFENSE: do NOT recommend legal escalation against the brand. Focus on monitoring + optional reputation reinforcement.

GATING RULES (avoid overreaction):
- If there are ZERO items with intent_framing="THREAT", then:
  - comms_package.external_holding_statement MUST be "not needed"
  - action_plan_next_4_hours MUST NOT include Legal unless an item explicitly indicates legal exposure for the brand.
- If intent_framing="DEFENSE", owner_team should be ["PR","Social"] or empty escalation; Legal only if brand is accused.
- Keep actions proportional to urgency (low=monitor/log, medium=prepare, high=activate crisis response).

OUTPUT RULES:
- Return ONLY valid JSON matching EXACTLY the schema below.
- Keep executive_summary to max 6 short bullets.
- In top_items_by_severity include max 5 items, each as a compact object with title, url, severity, intent_framing, urgency.
- In action_plan_next_4_hours include 3 to 6 actions maximum.
- Every action must reference a specific item_title from items_sample (or "cross-cutting").

REQUIRED JSON SCHEMA:
{{
  "ooda_timeline": {{
    "observe": "what was monitored and why",
    "orient": "how items were classified and scored",
    "decide": "how intent was determined and decisions made",
    "act": "what actions will be executed next"
  }},
  "executive_summary": ["max 6 bullets"],
  "situation_overview": {{
    "top_themes": ["..."],
    "overall_risk_level": "low|medium|high",
    "what_changed": "1-2 sentences",
    "why_now": "1-2 sentences"
  }},
  "decision_intelligence": {{
    "intent_distribution": {{}},
    "urgency_distribution": {{}},
    "top_items_by_severity": []
  }},
  "action_plan_next_4_hours": [
    {{
      "priority": 1,
      "item_title": "...",
      "intent_framing": "THREAT|DEFENSE|OPPORTUNITY|NEUTRAL|NOISE",
      "urgency": "low|medium|high",
      "objective": "...",
      "owner_team": ["PR","Legal","Security","Exec","Social"],
      "first_3_steps": ["...","...","..."],
      "success_criteria": ["..."],
      "notes": "short"
    }}
  ],
  "comms_package": {{
    "internal_message_draft": "short message to internal stakeholders",
    "external_holding_statement": "only if THREAT exists; otherwise 'not needed'",
    "optional_reinforcement_message": "useful especially for DEFENSE cases"
  }},
  "monitoring_and_triggers": {{
    "what_to_watch": ["..."],
    "update_frequency": "e.g. every 60 minutes",
    "escalation_triggers": ["..."],
    "de_escalation_triggers": ["..."]
  }},
  "risks_and_liability": {{
    "highest_risk_if_followed_blindly": "1-2 sentences",
    "human_judgment_overrides": ["..."]
  }}
}}

INPUT JSON (use as the ONLY source of truth):
{json.dumps(payload, ensure_ascii=False)}
""".strip()


def to_markdown(act: Dict, stats: Dict, items: List[Dict], brand: str, ts: str) -> str:
	lines = []
	lines.append(f"# ACT FULL — Executive Brief (OODA) — {brand}")
	lines.append(f"_Generated: {ts}_\n")

	lines.append("## OODA Timeline")
	lines.append("```json")
	lines.append(json.dumps(act.get("ooda_timeline", {}), ensure_ascii=False, indent=2))
	lines.append("```")

	lines.append("\n## Executive Summary")
	for b in act.get("executive_summary", []):
		lines.append(f"- {b}")

	lines.append("\n## Situation Overview")
	lines.append("```json")
	lines.append(json.dumps(act.get("situation_overview", {}), ensure_ascii=False, indent=2))
	lines.append("```")

	lines.append("\n## Decision Intelligence (distributions & top severity)")
	lines.append("```json")
	lines.append(json.dumps(act.get("decision_intelligence", {}), ensure_ascii=False, indent=2))
	lines.append("```")

	lines.append("\n## Action Plan (next 4 hours)")
	lines.append("```json")
	lines.append(json.dumps(act.get("action_plan_next_4_hours", []), ensure_ascii=False, indent=2))
	lines.append("```")

	lines.append("\n## Comms Package")
	lines.append("```json")
	lines.append(json.dumps(act.get("comms_package", {}), ensure_ascii=False, indent=2))
	lines.append("```")

	lines.append("\n## Monitoring & Triggers")
	lines.append("```json")
	lines.append(json.dumps(act.get("monitoring_and_triggers", {}), ensure_ascii=False, indent=2))
	lines.append("```")

	lines.append("\n## Risks & Liability")
	lines.append("```json")
	lines.append(json.dumps(act.get("risks_and_liability", {}), ensure_ascii=False, indent=2))
	lines.append("```")

	# --- Annex: STATS
	lines.append("\n# Annex A — Computed Stats")
	lines.append("```json")
	lines.append(json.dumps(stats, ensure_ascii=False, indent=2))
	lines.append("```")

	# --- Annex: ITEMS (full list)
	lines.append("\n# Annex B — Items Used (full)")
	lines.append("| decide_id | intent | urgency | severity | title | url |")
	lines.append("|---:|---|---|---:|---|---|")
	for it in items:
		lines.append(
			f"| {it.get('decide_id')} | {it.get('intent_framing')} | {it.get('urgency')} | {it.get('severity')} | "
			f"{(it.get('title') or '').replace('|',' ')} | {(it.get('url') or '')} |"
		)

	return "\n".join(lines)


def write_act_excel(path_xlsx: str, items: List[Dict]) -> None:
	wb = Workbook()
	ws = wb.active
	ws.title = "REPORT"

	headers = [
		"published_at",
		"title",
		"url",
		"severity",
		"intent_framing",
		"urgency",
		"narrative_category",
		"reputational_risk",
		"recommended_action",
		"snippet",
	]
	ws.append(headers)

	for cell in ws[1]:
		cell.font = Font(bold=True)

	for it in items:
		ws.append([
			it.get("published_at"),
			it.get("title"),
			it.get("url"),
			it.get("severity"),
			it.get("intent_framing"),
			it.get("urgency"),
			it.get("narrative_category"),
			it.get("reputational_risk"),
			it.get("recommended_action"),
			it.get("snippet"),
		])

	for r in range(2, ws.max_row + 1):
		url_cell = ws.cell(row=r, column=3)
		if url_cell.value:
			url_cell.hyperlink = url_cell.value
			url_cell.font = Font(color="0000FF", underline="single")

	widths = [18, 60, 60, 10, 14, 10, 18, 20, 35, 80]
	for i, w in enumerate(widths, start=1):
		ws.column_dimensions[get_column_letter(i)].width = w
	for row_cells in ws.iter_rows(min_row=2):
		for c in row_cells:
			c.alignment = Alignment(wrap_text=True, vertical="top")

	wb.save(path_xlsx)
	print("Excel report saved:", path_xlsx)


def main():
	load_dotenv()
	api_key = os.getenv("OPENAI_API_KEY", "").strip()
	if not api_key:
		raise RuntimeError("Missing OPENAI_API_KEY in .env")

	client = OpenAI(api_key=api_key)

	with open("config.yaml", "r", encoding="utf-8") as f:
		cfg = yaml.safe_load(f)

	brand = get_brand(cfg)
	db_path = cfg["storage"]["db_path"]

	items = fetch_full_ooda_view(db_path, brand, limit=50)
	if not items:
		print("No DECIDE items found. Run: python src/ooda_decide.py")
		return

	stats = compute_stats(items)

	# ACT via LLM (aggregated)
	prompt = build_act_prompt(brand, stats, items)

	resp = client.chat.completions.create(
		model=MODEL,
		messages=[{"role": "user", "content": prompt}],
		response_format={"type": "json_object"},
		timeout=90,
	)

	act_core = json.loads(resp.choices[0].message.content)

	# Build FULL payload (core + annex data)
	ts = datetime.now().strftime("%Y%m%d_%H%M%S")
	full = {
		"meta": {
			"project": cfg.get("project", {}),
			"generated_at": ts,
			"model": MODEL,
		},
		"act": act_core,
		"computed_stats": stats,
		"annex_data": {
			"items_used_full": items,
		},
	}

	# Save to DB
	conn = get_conn(db_path)
	ensure_act_table(conn)
	cur = conn.cursor()
	cur.execute(
		"INSERT INTO runs_act (brand, act_json) VALUES (?, ?)",
		(brand, json.dumps(full, ensure_ascii=False)),
	)
	conn.commit()
	conn.close()

	# Save outputs
	out_dir = os.getenv("RUN_DIR", "outputs")
	os.makedirs(out_dir, exist_ok=True)
	out_json = os.path.join(out_dir, f"act_full_{ts}.json")

	date_tag = datetime.now().strftime("%y%m%d_%H%M")
	out_xlsx = os.path.join(
		out_dir,
		f"AI brand reputation monitoring report {date_tag}.xlsx"
	)

	with open(out_json, "w", encoding="utf-8") as f:
		json.dump(full, f, ensure_ascii=False, indent=2)

	write_act_excel(out_xlsx, items)

	print("ACT FULL saved:")
	print("-", out_json)
	print("-", out_xlsx)


if __name__ == "__main__":
	main()
